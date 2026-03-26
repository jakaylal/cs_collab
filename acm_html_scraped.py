import os
import re
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook

# Use absolute paths
script_dir = os.path.dirname(os.path.abspath(__file__))
excel_file = os.path.join(script_dir, "acm_html_scraped_results.xlsx")

# Create Excel file if it doesn't exist
if not os.path.exists(excel_file):
    wb = Workbook()
    wb.save(excel_file)
else:
    try:
        wb = load_workbook(excel_file)
    except ValueError as e:
        print(f"Error loading Excel file: {e}. Creating a new one.")
        wb = Workbook()
        wb.save(excel_file)
        wb = load_workbook(excel_file)


if "Sheet1" not in wb.sheetnames:
    ws = wb.create_sheet("Sheet1")
else:
    ws = wb["Sheet1"]

start_row = 1006239
current_row = start_row

COLUMN_MAP = {
    "Year": "A",
    "Month": "B",
    "Letter": "C",
    "# of Search Results": "D",
    "Title": "E",
    "URL": "F",
    "Authors": "G",
    "Journal": "H",
    "Total Citations": "I",
    "Total Downloads": "J",
    "Last 12 months": "K",
    "Last 6 weeks": "L",
}


def get_year():
    query_span = soup.find("span", class_="query")
    if query_span:
        text = query_span.get_text()
        match = re.search(r"(\d{2}/\d{2}/(\d{4}))", text)
        if match:
            return match.group(2)
    return ""

def get_month():
    query_span = soup.find("span", class_="query")
    if query_span:
        text = query_span.get_text()
        match = re.search(r"(\d{2})/\d{2}/\d{4}", text)
        if match:
            return match.group(1)
    return ""

def get_letter():
    query_span = soup.find("span", class_="query")
    if query_span:
        text = query_span.get_text()
        match = re.search(r"\[Author:\s*([a-zA-Z])\]", text)
        if match:
            return match.group(1)
    return ""

def get_number_of_searches():
    pre_query = soup.find("span", class_="pre_query")
    if pre_query:
        # The number is typically before "Results" - look for digits
        parent = pre_query.find_parent()
        if parent:
            text = parent.get_text()
            match = re.search(r"(\d+)\s*Results", text)
            if match:
                return match.group(1)
    return ""

def get_title(context=None):
    ctx = context or soup
    title_link = ctx.find("h3", class_="issue-item__title")
    if title_link:
        a_tag = title_link.find("a")
        if a_tag:
            return a_tag.get_text(strip=True)
    return ""

def get_article_url(context=None):
    ctx = context or soup
    title_link = ctx.find("h3", class_="issue-item__title")
    if title_link:
        a_tag = title_link.find("a")
        if a_tag and a_tag.get("href"):
            href = a_tag.get("href")
            return href if href.startswith("http") else "https://dl.acm.org" + href
    return ""

def get_authors(context=None):
    ctx = context or soup
    author_links = []
    author_list = ctx.find("ul", class_="loa")
    if author_list:
        for li in author_list.find_all("li"):
            a_tag = li.find("a")
            if a_tag:
                name = a_tag.get_text(strip=True)
                href = a_tag.get("href", "")
                full_link = href if href.startswith("http") else "https://dl.acm.org" + href
                author_links.append(f"{name} ({full_link})")
    return " ||| ".join(author_links)


def get_total_citations(context=None):
    ctx = context or soup
    citation_div = ctx.find("div", class_="citation")
    if citation_div:
        span = citation_div.find("span", class_="bold")
        if span:
            return span.get_text(strip=True)
    return ""

def total_downloads(context=None):
    ctx = context or soup
    metric_div = ctx.find("div", class_="metric")
    if metric_div:
        span = metric_div.find("span", class_="bold")
        if span:
            return span.get_text(strip=True)
    return ""

def get_last_12_months(context=None):
    ctx = context or soup
    infos = ctx.find_all("div", class_="info")
    for info in infos:
        label_text = info.get_text()
        if "Last 12" in label_text or "12 Months" in label_text:
            span = info.find("span", class_="bold")
            if span:
                return span.get_text(strip=True)
    return ""

def get_last_6_weeks(context=None):
    ctx = context or soup
    infos = ctx.find_all("div", class_="info")
    for info in infos:
        label_text = info.get_text()
        if "Last 6" in label_text or "6 weeks" in label_text:
            span = info.find("span", class_="bold")
            if span:
                return span.get_text(strip=True)
    return ""

def get_journal(context=None):
    ctx = context or soup
    journal_link = ctx.find("a", href=lambda x: x and "/toc/" in x)
    if journal_link:
        span = journal_link.find("span", class_="epub-section__title")
        if span:
            return span.get_text(strip=True)
    return ""


html_folder = os.path.join(script_dir, "acm_html")
if not os.path.exists(html_folder):
    raise FileNotFoundError(f"HTML folder not found at: {html_folder}")

from string import ascii_uppercase

MAX_EXCEL_ROWS = 1048576  # Excel limit

sheet_index = 1
ws = wb.active
ws.title = f"Sheet{sheet_index}"

# Add headers if empty
if ws.max_row == 1:
    for header, col in COLUMN_MAP.items():
        ws[f"{col}1"] = header

current_row = ws.max_row + 1

for letter in ascii_uppercase:  # Letters A-Z
    for year in range(2009, 2012):  # Years 1990-2009
        for month in range(1, 13):  # Months 1-12
            file_name = os.path.join(html_folder, f"acm_{letter}_{year}_{month}_to_{year}_{month}_page_1.html")
            if not os.path.exists(file_name):
                continue

            with open(file_name, "r", encoding="utf-8") as f:
                soup = BeautifulSoup(f, "html.parser")

            # Page-level fields (same for all articles on the same page)
            page_year = get_year()
            page_month = get_month()
            page_letter = get_letter()
            page_num_results = get_number_of_searches()

            # Each search result has its own container
            articles = soup.find_all("li", class_="search__item issue-item-container")
            if not articles:
                # Fallback: if page structure differs, still try to scrape the first item
                articles = [soup]

            for article in articles:
                if current_row > MAX_EXCEL_ROWS:
                    sheet_index += 1
                    sheet_name = f"Sheet{sheet_index}"
                    ws = wb.create_sheet(title=sheet_name)

                    print(f"⚠️ Creating new sheet: {sheet_name}")

    # Write headers
                for header, col in COLUMN_MAP.items():
                    ws[f"{col}1"] = header

                current_row = 2

                ws[f"{COLUMN_MAP['Year']}{current_row}"] = page_year
                ws[f"{COLUMN_MAP['Month']}{current_row}"] = page_month
                ws[f"{COLUMN_MAP['Letter']}{current_row}"] = page_letter
                ws[f"{COLUMN_MAP['# of Search Results']}{current_row}"] = page_num_results
                ws[f"{COLUMN_MAP['Title']}{current_row}"] = get_title(article)
                ws[f"{COLUMN_MAP['URL']}{current_row}"] = get_article_url(article)
                ws[f"{COLUMN_MAP['Authors']}{current_row}"] = get_authors(article)
                ws[f"{COLUMN_MAP['Journal']}{current_row}"] = get_journal(article)
                ws[f"{COLUMN_MAP['Total Citations']}{current_row}"] = get_total_citations(article)
                ws[f"{COLUMN_MAP['Total Downloads']}{current_row}"] = total_downloads(article)
                ws[f"{COLUMN_MAP['Last 12 months']}{current_row}"] = get_last_12_months(article)
                ws[f"{COLUMN_MAP['Last 6 weeks']}{current_row}"] = get_last_6_weeks(article)

                current_row += 1

wb.save(excel_file)
print("Excel file has been saved")



